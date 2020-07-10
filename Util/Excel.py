from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
import time
from Projectar.var import *


class parseExcel(object):
    def __init__(self, excelPath):
        self.excelPath = excelPath
        self.workbook = load_workbook(excelPath)
        self.sheet = self.workbook.active
        self.font = Font(color=None)
        self.colorDict = {"red": 'FFFF3030', "green": 'FF008800'}

    def set_sheet_by_index(self, sheet_index):
        sheet_name = self.workbook.get_sheet_names()[sheet_index]
        self.sheet = self.workbook.get_sheet_by_name(sheet_name)
        return self.sheet

    def get_default_sheet(self):
        return self.sheet.title

    def set_sheet_by_name(self, sheet_name):
        sheet = self.workbook.get_sheet_by_name(sheet_name)
        self.sheet = sheet
        return self.sheet

    def get_max_row_no(self):
        return self.sheet.max_row

    def get_min_row_no(self):
        return self.sheet.min_row

    def get_max_col_no(self):
        return self.sheet.max_column

    def get_min_col_no(self):
        return self.sheet.min_column

    def get_all_rows(self):
        return list(self.sheet.iter_rows())

    def get_all_cols(self):
        return list(self.sheet.iter_cols())

    def get_single_col(self, col_no):
        return self.get_all_cols()[col_no]

    def get_single_row(self, row_no):
        return self.get_all_rows()[row_no]

    def get_cell(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no)

    def get_cell_content(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no).value

    def write_cell_content(self, row_no, col_no, content, font=None):
        self.sheet.cell(row=row_no, column=col_no).value = content
        self.workbook.save(self.excelPath)
        return self.sheet.cell(row=row_no, column=row_no).value

    def write_cell_current_time(self, row_no, col_no):
        time1 = time.strftime("%Y-%m-%d %H:%M:%S")
        self.sheet.cell(row=row_no, column=col_no).value = str(time1)
        self.workbook.save(self.excelPath)
        return self.sheet.cell(row=row_no, column=row_no).value

    def save_excel_file(self):
        self.workbook.save(self.excelPath)


if __name__ == '__main__':
    p = parseExcel("C:\\Users\\v_lvjichao\\PycharmProjects\\dataDrivenTestPractice\\TestData\\126mail.xlsx")
    print(p.get_default_sheet())
    print(p.set_sheet_by_index(0))
    print(p.get_default_sheet())
    print(p.set_sheet_by_index(1))
    print(p.get_default_sheet())
    print(p.get_max_row_no())
    print(p.get_max_col_no())
    print(p.get_min_col_no())
    print(p.get_min_row_no())
    print(p.get_all_rows())
    print(p.get_all_cols())
    print(p.get_single_col(2))
    print(p.get_single_row(1))
    print(p.get_cell(2, 2))
    print(p.get_cell_content(2, 2))
    print(p.write_cell_content(11, 11, "xiaoxiaoyu"))
    print(p.write_cell_current_time(13, 13))
